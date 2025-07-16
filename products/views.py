# C:\Users\DELL\SER SQL MY APP\products\views.py

from decimal import Decimal
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.decorators import action
from rest_framework import serializers
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import F, Q, Sum # Added Sum for aggregation
from django.conf import settings
import openpyxl
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.exceptions import ValidationError, NotFound # Added NotFound
# from django.contrib import messages # Import messages for ProductUploadExcelView warnings - Not needed for API response

# استيراد النماذج الصحيحة
from .models import Department, Product, ProductCategory, BranchProductInventory, InventoryMovement
from .serializers import DepartmentSerializer, ProductSerializer, ProductCategorySerializer, BranchProductInventorySerializer
from users.models import UserAccount, Role
from stores.models import Branch
from geopy.distance import geodesic # مكتبة لحساب المسافة الجغرافية
from django.utils.translation import gettext_lazy as _

# استيراد CustomPermission
from mysite.permissions import CustomPermission


# --- API ViewSets for Product Categories ---
class ProductCategoryViewSet(viewsets.ModelViewSet):
    queryset = ProductCategory.objects.all()
    serializer_class = ProductCategorySerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated, CustomPermission]

    def get_queryset(self):
        user = self.request.user
        # مالك التطبيق / المشرف العام / مدير المشروع / فريق عمل التطبيق يمكنهم رؤية جميع الفئات
        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            return ProductCategory.objects.all()
        # العملاء، مدراء المتاجر والفروع، الموظفون العامون، الكاشير، منظمو الرفوف، موظفو خدمة العملاء يمكنهم رؤية جميع الفئات (للقراءة فقط)
        if user.is_platform_customer() or user.is_store_manager_user() or \
           user.is_branch_manager_user() or user.is_general_staff_user() or \
           user.is_cashier_user() or user.is_shelf_organizer_user() or user.is_customer_service_user():
            return ProductCategory.objects.all()
        return ProductCategory.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if not (user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()):
            raise ValidationError({'detail': _('You do not have permission to create product categories.')})
        serializer.save()

    def perform_update(self, serializer):
        user = self.request.user
        if not (user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()):
            raise ValidationError({'detail': _('You do not have permission to update product categories.')})
        serializer.save()

    def perform_destroy(self, instance):
        user = self.request.user
        if not (user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()):
            raise ValidationError({'detail': _('You do not have permission to delete product categories.')})
        instance.delete()


# --- API ViewSets for Departments ---
class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated, CustomPermission]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            return Department.objects.all()
        
        if user.is_store_manager_user() and user.store:
            return Department.objects.filter(branch__store=user.store)
        
        if (user.is_branch_manager_user() or user.is_general_staff_user() or \
            user.is_cashier_user() or user.is_shelf_organizer_user() or user.is_customer_service_user()) and user.branch:
            return Department.objects.filter(branch=user.branch)
            
        if user.is_platform_customer():
            return Department.objects.all() # العملاء يمكنهم رؤية جميع الأقسام (للقراءة فقط)
            
        return Department.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        branch = serializer.validated_data.get('branch')

        if not branch:
            raise ValidationError({'branch': _('Branch must be specified for the department.')})

        if user.is_superuser or user.is_app_owner() or user.is_project_manager():
            pass # يمكنهم إنشاء أي قسم في أي فرع
        elif user.is_store_manager_user() and user.store and user.store == branch.store:
            pass # مدير المتجر يمكنه إنشاء قسم في أي فرع ضمن متجره
        elif user.is_branch_manager_user() and user.branch and user.branch == branch:
            pass # مدير الفرع يمكنه إنشاء قسم في فرعه فقط
        else:
            raise ValidationError({'error': _('You do not have permission to create a department in this branch.')})

        serializer.save() # إزالة created_by=user حيث Department model لا يملك هذا الحقل

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object() 
        new_branch = serializer.validated_data.get('branch', instance.branch) # احصل على الفرع الجديد أو الحالي

        if user.is_superuser or user.is_app_owner() or user.is_project_manager():
            pass # يمكنهم تعديل أي قسم
        elif user.is_store_manager_user() and user.store and user.store == instance.branch.store:
            if new_branch != instance.branch and user.store != new_branch.store: # منع نقل قسم خارج المتجر الخاص به
                raise ValidationError({'error': _('You cannot move a department to a branch outside your store.')})
        elif user.is_branch_manager_user() and user.branch and user.branch == instance.branch:
            if new_branch != instance.branch: # منع نقل قسم من الفرع الخاص به
                raise ValidationError({'error': _('You cannot move a department from your branch to another branch.')})
        else:
            raise ValidationError({'error': _('You do not have permission to modify this department.')})
            
        serializer.save() # إزالة last_updated_by=user حيث Department model لا يملك هذا الحقل

    def perform_destroy(self, instance):
        user = self.request.user
        if user.is_superuser or user.is_app_owner() or user.is_project_manager():
            instance.delete()
        elif user.is_store_manager_user() and user.store and user.store == instance.branch.store:
            instance.delete()
        elif user.is_branch_manager_user() and user.branch and user.branch == instance.branch:
            instance.delete()
        else:
            raise ValidationError({'error': _('You do not have permission to delete this department.')})


# --- API ViewSets for Products ---
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated, CustomPermission]

    def get_queryset(self):
        user = self.request.user
        base_qs = Product.objects.all()

        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            return base_qs
            
        # مدراء المتاجر يرون المنتجات الموجودة في مخزون فروع متجرهم
        if user.is_store_manager_user() and user.store:
            return base_qs.filter(branch_inventories__branch__store=user.store).distinct()
            
        # مدراء الفروع أو الموظفون العامون يرون المنتجات الموجودة في مخزون فرعهم
        if (user.is_branch_manager_user() or user.is_general_staff_user() or \
            user.is_cashier_user() or user.is_shelf_organizer_user() or user.is_customer_service_user()) and user.branch:
            return base_qs.filter(branch_inventories__branch=user.branch).distinct()
            
        # العملاء يرون جميع المنتجات (فقط للقراءة)
        if user.is_platform_customer():
            return base_qs # يمكنهم رؤية جميع المنتجات المتاحة في النظام

        return Product.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        
        # صلاحية إنشاء المنتجات:
        # مالك التطبيق/المشرف العام/مدير المشروع/فريق عمل التطبيق يمكنهم إنشاء منتجات
        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            product = serializer.save(last_updated_by=user)
            return product
        else:
            raise ValidationError({'error': _('You do not have permission to create products.')})

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object() # المنتج الذي يتم تعديله
        
        # صلاحية تعديل المنتجات:
        # مالك التطبيق/المشرف العام/مدير المشروع/فريق عمل التطبيق يمكنهم تعديل أي منتج
        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            serializer.save(last_updated_by=user)
        # مدراء المتاجر يمكنهم تعديل المنتجات التي تقع ضمن نطاق متجرهم
        elif user.is_store_manager_user() and user.store and instance.branch_inventories.filter(branch__store=user.store).exists():
            serializer.save(last_updated_by=user)
        # مدراء الفروع يمكنهم تعديل المنتجات التي تقع ضمن نطاق فرعهم
        elif user.is_branch_manager_user() and user.branch and instance.branch_inventories.filter(branch=user.branch).exists():
            serializer.save(last_updated_by=user)
        else:
            raise ValidationError({'error': _('You do not have permission to modify this product.')})
            
    def perform_destroy(self, instance):
        user = self.request.user
        # صلاحية حذف المنتجات:
        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            instance.delete()
        elif user.is_store_manager_user() and user.store and instance.branch_inventories.filter(branch__store=user.store).exists():
            instance.delete()
        elif user.is_branch_manager_user() and user.branch and instance.branch_inventories.filter(branch=user.branch).exists():
            instance.delete()
        else:
            raise ValidationError({'error': _('You do not have permission to delete this product.')})

# --- API ViewSets for Branch Product Inventory ---
class BranchProductInventoryViewSet(viewsets.ModelViewSet):
    queryset = BranchProductInventory.objects.all()
    serializer_class = BranchProductInventorySerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated, CustomPermission]

    def get_queryset(self):
        user = self.request.user
        qs = BranchProductInventory.objects.all()

        if user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user():
            return qs
            
        # مدير المتجر يرى مخزون المنتجات في فروع متجره
        if user.is_store_manager_user() and user.store:
            return qs.filter(branch__store=user.store)
            
        # مدير الفرع أو الموظفون العامون أو الكاشير أو منظمو الرفوف أو موظفو خدمة العملاء يرون مخزون المنتجات في فرعهم
        if (user.is_branch_manager_user() or user.is_general_staff_user() or \
            user.is_cashier_user() or user.is_shelf_organizer_user() or user.is_customer_service_user()) and user.branch:
            return qs.filter(branch=user.branch)
            
        # العملاء لا يرون سجلات المخزون (أو يرون فقط المنتجات المتاحة من خلال ProductViewSet)
        if user.is_platform_customer():
            return BranchProductInventory.objects.none()
            
        return BranchProductInventory.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        product = serializer.validated_data.get('product')
        branch = serializer.validated_data.get('branch')

        # يجب أن يكون لدى المستخدم صلاحية لإنشاء سجل مخزون للمنتج في الفرع المحدد
        if not (user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()):
            if not (user.is_store_manager_user() and user.store and user.store == branch.store) and \
               not (user.is_branch_manager_user() and user.branch and user.branch == branch) and \
               not (user.is_shelf_organizer_user() and user.department and user.department.branch == branch): # منظم الرفوف يمكنه إنشاء مخزون لمنتجات في قسمه بفرعه
                raise ValidationError({'detail': _('You do not have permission to create inventory for this branch.')})

        # تحقق من وجود سجل مخزون لهذا المنتج في هذا الفرع بالفعل
        if BranchProductInventory.objects.filter(product=product, branch=branch).exists():
            raise ValidationError({'detail': _('Inventory record for this product in this branch already exists. Please update it instead.')})

        serializer.save(last_updated_by=user)

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object() # سجل المخزون الذي يتم تعديله
        
        # يجب أن يكون لدى المستخدم صلاحية لتعديل سجل المخزون هذا
        if not (user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()):
            if not (user.is_store_manager_user() and user.store and user.store == instance.branch.store) and \
               not (user.is_branch_manager_user() and user.branch and user.branch == instance.branch) and \
               not (user.is_general_staff_user() and user.branch and user.branch == instance.branch) and \
               not (user.is_shelf_organizer_user() and user.department and user.department.branch == instance.branch and instance.product.department == user.department):
                raise ValidationError({'detail': _('You do not have permission to update this inventory record.')})
        
        serializer.save(last_updated_by=user)

    def perform_destroy(self, instance):
        user = self.request.user
        # صلاحية حذف سجلات المخزون:
        if not (user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()):
            raise ValidationError({'detail': _('You do not have permission to delete inventory records.')})
        instance.delete()


# --- Product Excel Upload View ---
class ProductUploadExcelView(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated, CustomPermission]

    @action(detail=False, methods=['post'])
    def upload(self, request):
        user = request.user
        
        if not (user.is_store_manager_user() or user.is_branch_manager_user() or user.is_superuser or user.is_app_owner() or user.is_project_manager() or user.is_app_staff_user()): # Added app_staff_user
            return Response({'error': _('You do not have permission to upload product Excel files.')}, status=status.HTTP_403_FORBIDDEN)

        file = request.data.get('file')
        branch_id = request.data.get('branch_id')

        if not file:
            return Response({'detail': _('No file uploaded.')}, status=status.HTTP_400_BAD_REQUEST)
            
        if not branch_id:
            return Response({'detail': _('Branch ID is required for product upload.')}, status=status.HTTP_400_BAD_REQUEST)

        try:
            branch = Branch.objects.get(id=branch_id)
            # التحقق من أن المستخدم لديه صلاحية على هذا الفرع الذي يحاول الرفع إليه
            if user.is_store_manager_user() and user.store and user.store != branch.store:
                return Response({'detail': _('You do not have permission to upload products to this store\'s branches.')}, status=status.HTTP_403_FORBIDDEN)
            if user.is_branch_manager_user() and user.branch and user.branch != branch:
                return Response({'detail': _('You do not have permission to upload products to this branch.')}, status=status.HTTP_403_FORBIDDEN)
            # Added for app staff
            if user.is_app_staff_user() and not (user.is_superuser or user.is_app_owner() or user.is_project_manager()):
                # App staff can upload to any branch if they are superuser/app owner/project manager, otherwise not.
                # If they are just 'app_staff_user' without other high roles, they can't upload.
                return Response({'detail': _('App Staff users without higher privileges cannot upload product Excel files to specific branches.')}, status=status.HTTP_403_FORBIDDEN)


        except Branch.DoesNotExist:
            return Response({'detail': _('Branch not found.')}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            header = [cell.value for cell in sheet[1]]
            
            # Column names for Product model
            barcode_col_name = _("Barcode")
            item_number_col_name = _("Item Number")
            name_col_name = _("Product Name")
            price_col_name = _("Base Price")
            department_col_name = _("Department Name")
            category_col_name = _("Category Name") # New column for Category
            vat_rate_col_name = _("VAT Rate (as decimal)")

            # Column name for BranchProductInventory
            quantity_col_name = _("Quantity in Stock") # Quantity is now in BranchProductInventory

            products_to_create_count = 0
            products_to_update_count = 0
            inventory_records_processed_count = 0
            
            # Using a list to collect error messages from rows
            row_errors = []

            with transaction.atomic():
                for row_idx in range(2, sheet.max_row + 1):
                    row_data = {header[col_idx]: sheet.cell(row=row_idx, column=col_idx + 1).value
                                for col_idx in range(len(header))}
                    
                    barcode = str(row_data.get(barcode_col_name)).strip() if row_data.get(barcode_col_name) else None
                    item_number = str(row_data.get(item_number_col_name)).strip() if row_data.get(item_number_col_name) else None
                    name = str(row_data.get(name_col_name)).strip() if row_data.get(name_col_name) else None
                    price = row_data.get(price_col_name)
                    quantity = row_data.get(quantity_col_name) # This is for BranchProductInventory
                    department_name = str(row_data.get(department_col_name)).strip() if row_data.get(department_col_name) else None
                    category_name = str(row_data.get(category_col_name)).strip() if row_data.get(category_col_name) else None
                    vat_rate = row_data.get(vat_rate_col_name)

                    if not name or price is None: # Price can be 0, so check for None
                        row_errors.append(_(f"Row {row_idx}: Product Name or Price missing. Skipping."))
                        continue
                    
                    if not (barcode or item_number):
                        row_errors.append(_(f"Row {row_idx}: Barcode or Item Number is required for product identification. Skipping."))
                        continue

                    try:
                        price = Decimal(str(price))
                    except (TypeError, ValueError):
                        row_errors.append(_(f"Row {row_idx}: Invalid price format. Skipping."))
                        continue

                    try:
                        vat_rate = Decimal(str(vat_rate)) if vat_rate is not None else Decimal('0.1500')
                        if not (Decimal('0.0000') <= vat_rate <= Decimal('1.0000')):
                             raise ValueError("VAT rate out of range.")
                    except (TypeError, ValueError):
                        row_errors.append(_(f"Row {row_idx}: Invalid VAT rate format. Using default 0.15."))
                        vat_rate = Decimal('0.1500')

                    # Find or create Department
                    department_obj = None
                    if department_name:
                        # Ensure department belongs to the branch
                        department_obj, created = Department.objects.get_or_create(
                            branch=branch,
                            name=department_name
                        )
                    
                    # Find or create ProductCategory
                    category_obj = None
                    if category_name:
                        category_obj, created = ProductCategory.objects.get_or_create(
                            name=category_name
                        )

                    # Find existing product (globally, not tied to a specific branch in Product model)
                    existing_product = None
                    if barcode:
                        existing_product = Product.objects.filter(barcode=barcode).first()
                    if not existing_product and item_number:
                        existing_product = Product.objects.filter(item_number=item_number).first()

                    product_data = {
                        'name': name,
                        'price': price,
                        'department': department_obj,
                        'category': category_obj, # Assign category
                        'last_updated_by': user,
                        'vat_rate': vat_rate
                    }
                    if barcode: product_data['barcode'] = barcode
                    if item_number: product_data['item_number'] = item_number

                    current_product = None
                    if existing_product:
                        # Update existing product (Product model attributes)
                        for key, value in product_data.items():
                            setattr(existing_product, key, value)
                        existing_product.save(update_fields=list(product_data.keys())) # Update specific fields
                        current_product = existing_product
                        products_to_update_count += 1
                    else:
                        # Create new product
                        current_product = Product.objects.create(**product_data)
                        products_to_create_count += 1
                    
                    # Handle BranchProductInventory for the current product and branch
                    if current_product and quantity is not None:
                        try:
                            inventory_quantity = int(quantity)
                            if inventory_quantity < 0:
                                raise ValueError("Quantity cannot be negative.")
                        except (TypeError, ValueError):
                            row_errors.append(_(f"Row {row_idx}: Invalid quantity format for product '{name}'. Using 0."))
                            inventory_quantity = 0
                            
                        branch_product_inventory, created_inventory = BranchProductInventory.objects.get_or_create(
                            product=current_product,
                            branch=branch,
                            defaults={
                                'quantity': inventory_quantity,
                                'last_updated_by': user
                            }
                        )
                        if not created_inventory:
                            # Update existing inventory
                            branch_product_inventory.quantity = inventory_quantity
                            branch_product_inventory.last_updated_by = user
                            branch_product_inventory.save(update_fields=['quantity', 'last_updated_by'])
                        
                        inventory_records_processed_count += 1

                # If any row errors occurred, return them with a partial success or warning status
                if row_errors:
                    return Response({
                        'detail': _('Products and inventory processed with some warnings/errors.'),
                        'created_products_count': products_to_create_count,
                        'updated_products_count': products_to_update_count,
                        'inventory_records_processed': inventory_records_processed_count,
                        'row_errors': row_errors
                    }, status=status.HTTP_200_OK) # Or HTTP_206_PARTIAL_CONTENT
                
                return Response({
                    'detail': _('Products and inventory uploaded successfully.'),
                    'created_products_count': products_to_create_count,
                    'updated_products_count': products_to_update_count,
                    'inventory_records_processed': inventory_records_processed_count
                }, status=status.HTTP_200_OK)

        except Exception as e:
            # Catch any unexpected errors during file processing
            return Response({'detail': _(f'Error processing file: {e}')}, status=status.HTTP_400_BAD_REQUEST)


# --- API View for Barcode Scanning (Strict Location Check) ---
class ScanBarcodeAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        barcode_value = request.data.get('barcode')
        customer_lat = request.data.get('customer_latitude')
        customer_lon = request.data.get('customer_longitude')

        if not barcode_value:
            return Response({"detail": _("Barcode is required.")}, status=status.HTTP_400_BAD_REQUEST)
            
        if customer_lat is None or customer_lon is None:
            return Response({"detail": _("Customer location (latitude, longitude) is required to scan products.")}, status=status.HTTP_400_BAD_REQUEST)

        try:
            customer_coords = (float(customer_lat), float(customer_lon))
        except (TypeError, ValueError):
            return Response({"detail": _("Invalid latitude or longitude format for customer location.")}, status=status.HTTP_400_BAD_REQUEST)

        MAX_DISTANCE_FOR_SCAN = getattr(settings, 'MAX_BARCODE_SCAN_DISTANCE_KM', 0.01) # Default: 10 meters

        branches_with_coords = Branch.objects.filter(latitude__isnull=False, longitude__isnull=False)
        
        if not branches_with_coords.exists():
            return Response({"detail": _("No branches with valid GPS coordinates are configured. Cannot proceed with barcode scanning.")}, status=status.HTTP_404_NOT_FOUND)

        detected_branch = None
        for branch in branches_with_coords:
            branch_coords = (branch.latitude, branch.longitude)
            try:
                distance = geodesic(customer_coords, branch_coords).km
                if distance <= MAX_DISTANCE_FOR_SCAN:
                    detected_branch = branch
                    break # Found a nearby branch, no need to check others
            except ValueError:
                # Handle cases where geodesic might fail due to invalid coordinates (though already checked for floats)
                continue

        if not detected_branch:
            return Response(
                {"detail": _(f"You are currently not inside any store's scanning zone (max {MAX_DISTANCE_FOR_SCAN * 1000} meters from entrance). Barcode scanning is restricted to in-store use.")},
                status=status.HTTP_403_FORBIDDEN
            )

        # 3. If branch is detected, find the Product and its inventory in that branch
        try:
            product = Product.objects.get(barcode=barcode_value) # Find product globally by barcode
            
            # Now, get the specific BranchProductInventory for this product in the detected branch
            product_inventory = BranchProductInventory.objects.get(product=product, branch=detected_branch)

            # Return Product data along with its quantity in the detected branch
            serializer = ProductSerializer(product)
            
            response_data = serializer.data
            
            # Find the specific inventory data for the detected branch
            current_branch_inventory_data = next((inv for inv in response_data.get('branch_inventories', []) if inv['branch'] == detected_branch.id), None)
            
            if current_branch_inventory_data:
                response_data['quantity_in_scanned_branch'] = current_branch_inventory_data['quantity']
                response_data['scanned_branch_name'] = detected_branch.name # Add scanned branch name
            else:
                response_data['quantity_in_scanned_branch'] = 0 # Product exists, but no inventory in this specific branch
                response_data['scanned_branch_name'] = detected_branch.name

            # Optionally remove the full list of branch_inventories if you only need the scanned branch's quantity
            if 'branch_inventories' in response_data:
                del response_data['branch_inventories']

            return Response(response_data, status=status.HTTP_200_OK)
        except Product.DoesNotExist:
            return Response(
                {"detail": _(f"Product with barcode '{barcode_value}' not found.")},
                status=status.HTTP_404_NOT_FOUND
            )
        except BranchProductInventory.DoesNotExist:
            # If product exists but no inventory in the detected branch
            product = Product.objects.get(barcode=barcode_value) # Re-fetch product to get its details
            return Response(
                {"detail": _(f"Product '{product.name}' found, but no inventory record exists for it in branch '{detected_branch.name}'."),
                 "product_details": ProductSerializer(product).data, # Optionally return product details without inventory
                 "quantity_in_scanned_branch": 0,
                 "scanned_branch_name": detected_branch.name
                },
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

